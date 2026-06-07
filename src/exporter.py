from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COLUNAS_FINAIS = [
    "data_referencia",
    "moeda_iso",
    "cotacao_b3",
    "ptax_compra",
    "ptax_venda",
    "spread_b3_ptax",
    "instrumento_b3",
    "metodo_calculo",
]

COLUNAS_TAXA = {"cotacao_b3", "ptax_compra", "ptax_venda"}


def gerar_excel(df: pd.DataFrame, data_ref: date, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    caminho = output_dir / f"taxas_spot_b3_{data_ref.strftime('%Y%m%d')}.xlsx"

    df = df.copy()
    df.insert(0, "data_referencia", data_ref.isoformat())
    df = df[COLUNAS_FINAIS]

    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Taxas B3 vs PTAX", index=False)

    wb = load_workbook(caminho)
    ws = wb["Taxas B3 vs PTAX"]

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    hdr_align = Alignment(horizontal="center")
    for cell in ws[1]:
        cell.font, cell.fill, cell.alignment = hdr_font, hdr_fill, hdr_align

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            col_name = ws.cell(row=1, column=cell.column).value
            if col_name in COLUNAS_TAXA:
                cell.number_format = "#,##0.0000"
            elif col_name == "spread_b3_ptax":
                cell.number_format = "0.0000"

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max_len + 4, 30
        )

    ws.freeze_panes = "A2"
    wb.save(caminho)
    return caminho
